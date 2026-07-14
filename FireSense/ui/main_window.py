"""
main_window.py — FireSense main window.

Pages:
  Setup     — user configures their own OPNsense connection (host, API keys, …)
  Deploy    — pick which deployment stages to run, validate, then start
  Dashboard — stat cards + live Q-values + action distribution + recent table
  Monitor   — full live Q-values chart + all decisions
  Logs      — colored system log
"""

import os
import sys
import csv
import time
import shutil
import collections
import datetime

import pyqtgraph as pg
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QHBoxLayout, QVBoxLayout, QGridLayout, QLabel,
    QPushButton, QTableWidget, QTableWidgetItem, QPlainTextEdit, QFrame,
    QHeaderView, QStackedWidget, QScrollArea, QLineEdit, QCheckBox, QSpinBox,
    QAbstractSpinBox, QMessageBox, QFileDialog,
)
from PyQt6.QtCore import Qt, pyqtSignal, QTimer
from PyQt6.QtGui  import QPixmap, QColor, QIcon, QGuiApplication

from .styles import (
    DARK_THEME, ACTION_COLORS, Q_COLORS, Q_NAMES, STATUS_COLORS,
    ORANGE, ORANGE_HI, AMBER, GREEN, BLUE, RED, MUTED, FAINT, TEXT, CARD, BG0,
    BG1, BORDER2,
)
from config_manager import ConfigManager
from stages import StageRunner, STAGE_ORDER, STAGE_META, STAGE_FUNCS

pg.setConfigOption("background", BG0)
pg.setConfigOption("foreground", MUTED)
pg.setConfigOption("antialias", True)

HISTORY = 30


# ────────────────────────────────────────────────────────────────────────────
# Small reusable widgets
# ────────────────────────────────────────────────────────────────────────────

class SplashScreen(QWidget):
    """Frameless, centered startup splash: big logo + wordmark + status line.
    Shown while the (heavier) main window and model resources initialize."""
    def __init__(self, logo_path):
        super().__init__(None,
            Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint |
            Qt.WindowType.SplashScreen)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setFixedSize(420, 380)

        card = QWidget(self); card.setObjectName("splash")
        card.setGeometry(0, 0, 420, 380)
        lay = QVBoxLayout(card)
        lay.setContentsMargins(30, 40, 30, 30); lay.setSpacing(6)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)

        logo = QLabel()
        logo.setPixmap(QPixmap(logo_path).scaled(150, 150,
            Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
        logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title = QLabel("FireSense"); title.setObjectName("splash_title")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub = QLabel("DQN FIREWALL"); sub.setObjectName("splash_sub")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._status = QLabel("Secure with Sense."); self._status.setObjectName("splash_status")
        self._status.setAlignment(Qt.AlignmentFlag.AlignCenter)

        lay.addStretch()
        lay.addWidget(logo); lay.addSpacing(14)
        lay.addWidget(title); lay.addWidget(sub)
        lay.addStretch()
        lay.addWidget(self._status)
        self.setStyleSheet(DARK_THEME)

    def set_status(self, msg):
        self._status.setText(msg)

    def center_on_screen(self):
        screen = self.screen() or QGuiApplication.primaryScreen()
        if screen:
            geo = screen.availableGeometry()
            self.move(geo.center().x() - self.width() // 2,
                      geo.center().y() - self.height() // 2)


class Divider(QFrame):
    def __init__(self, vertical=False):
        super().__init__()
        self.setFrameShape(QFrame.Shape.VLine if vertical else QFrame.Shape.HLine)


class ElidedLabel(QLabel):
    """QLabel that truncates with '…' instead of clipping when space is tight."""
    def __init__(self, text=""):
        self._full = text
        super().__init__()
        QLabel.setText(self, text)

    def setText(self, text):
        self._full = text
        self._apply_elide()

    def resizeEvent(self, e):
        super().resizeEvent(e)
        self._apply_elide()

    def _apply_elide(self):
        elided = self.fontMetrics().elidedText(
            self._full, Qt.TextElideMode.ElideRight, max(self.width(), 0))
        QLabel.setText(self, elided)
        self.setToolTip(self._full)


def _table_with_empty(table, message):
    """Wrap a QTableWidget in a QStackedWidget: page 0 = friendly empty state,
    page 1 = the table. Call stack.setCurrentIndex(...) to toggle."""
    stack = QStackedWidget()
    empty = QLabel(message); empty.setObjectName("empty_state")
    empty.setAlignment(Qt.AlignmentFlag.AlignCenter); empty.setWordWrap(True)
    holder = QWidget(); hl = QVBoxLayout(holder)
    hl.setContentsMargins(0, 0, 0, 0); hl.addStretch(); hl.addWidget(empty); hl.addStretch()
    stack.addWidget(holder)   # index 0
    stack.addWidget(table)    # index 1
    stack.setCurrentIndex(0)
    return stack


class StatCard(QWidget):
    def __init__(self, title, value="—", color="#ffffff", hint="", icon=""):
        super().__init__()
        self.setObjectName("card")
        lay = QVBoxLayout(self); lay.setContentsMargins(16, 14, 16, 14); lay.setSpacing(4)

        head = QHBoxLayout(); head.setContentsMargins(0, 0, 0, 0); head.setSpacing(7)
        if icon:
            ic = QLabel(icon); ic.setObjectName("card_icon")
            ic.setStyleSheet(f"color:{color};")
            head.addWidget(ic)
        self._t = QLabel(title.upper()); self._t.setObjectName("card_title")
        head.addWidget(self._t); head.addStretch()
        lay.addLayout(head)

        self._v = QLabel(value); self._v.setObjectName("card_value")
        self._v.setStyleSheet(f"color:{color};")
        self._h = QLabel(hint); self._h.setObjectName("card_hint")
        lay.addWidget(self._v); lay.addWidget(self._h)

    def set_value(self, v, color=None):
        self._v.setText(v)
        if color: self._v.setStyleSheet(f"color:{color};")

    def set_hint(self, h): self._h.setText(h)


class Field(QWidget):
    """Labeled input row used by the Setup form."""
    def __init__(self, label, widget):
        super().__init__()
        lay = QVBoxLayout(self); lay.setContentsMargins(0, 0, 0, 0); lay.setSpacing(5)
        lbl = QLabel(label.upper()); lbl.setObjectName("field_label")
        lay.addWidget(lbl); lay.addWidget(widget)
        self.widget = widget


class QValueChart(pg.PlotWidget):
    def __init__(self, height=220):
        super().__init__()
        self.setLabel("left", "Q-Value"); self.setLabel("bottom", "Jendela")
        self.addLegend(offset=(10, 8)); self.showGrid(x=True, y=True, alpha=0.12)
        self.setMinimumHeight(height)
        self._x = collections.deque(maxlen=HISTORY)
        self._d = [collections.deque(maxlen=HISTORY) for _ in range(4)]
        self._curves = []
        for color, name in zip(Q_COLORS, Q_NAMES):
            c = self.plot([], [], pen=pg.mkPen(color, width=2), name=name,
                          symbol="o", symbolSize=5, symbolBrush=color, symbolPen=None)
            self._curves.append(c)

    def push(self, wid, q):
        self._x.append(wid)
        for i in range(4):
            self._d[i].append(q[i] if i < len(q) else 0)
        x = list(self._x)
        for i, c in enumerate(self._curves):
            c.setData(x, list(self._d[i]))

    def reset(self):
        self._x.clear()
        for d in self._d: d.clear()
        for c in self._curves: c.setData([], [])


class ActionDistChart(pg.PlotWidget):
    ACTIONS = ["maintain", "block_ip", "tighten", "rollback"]
    def __init__(self):
        super().__init__()
        self.setLabel("left", "Count"); self.showGrid(y=True, alpha=0.12)
        self.setMinimumHeight(170)
        self.getAxis("bottom").setTicks([[(i, a) for i, a in enumerate(self.ACTIONS)]])
        self._counts = [0, 0, 0, 0]
        cols = [ACTION_COLORS[a] for a in self.ACTIONS]
        self._bars = pg.BarGraphItem(x=list(range(4)), height=self._counts, width=0.6,
                                     brushes=[pg.mkBrush(c) for c in cols],
                                     pens=[pg.mkPen(c, width=1.4) for c in cols])
        self.addItem(self._bars)

    def increment(self, a, n=1):
        if 0 <= a <= 3 and n > 0:
            self._counts[a] += n
            self._bars.setOpts(height=list(self._counts))

    def reset(self):
        self._counts = [0, 0, 0, 0]
        self._bars.setOpts(height=list(self._counts))


# ────────────────────────────────────────────────────────────────────────────
# Sidebar
# ────────────────────────────────────────────────────────────────────────────

class Sidebar(QWidget):
    def __init__(self, logo_path):
        super().__init__()
        self.setObjectName("sidebar")
        lay = QVBoxLayout(self); lay.setContentsMargins(0, 0, 0, 0); lay.setSpacing(0)

        top = QWidget(); tl = QVBoxLayout(top)
        tl.setContentsMargins(18, 24, 18, 16); tl.setSpacing(2)
        tl.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        logo = QLabel()
        logo.setPixmap(QPixmap(logo_path).scaled(72, 72, Qt.AspectRatioMode.KeepAspectRatio,
                                                 Qt.TransformationMode.SmoothTransformation))
        logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title = QLabel("FireSense"); title.setObjectName("app_title")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub = QLabel("DQN FIREWALL"); sub.setObjectName("app_subtitle")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        tl.addWidget(logo); tl.addSpacing(8); tl.addWidget(title); tl.addWidget(sub)
        lay.addWidget(top); lay.addWidget(Divider())

        nav = QWidget(); nav.setObjectName("sidebar_nav")
        nl = QVBoxLayout(nav); nl.setContentsMargins(0, 8, 0, 0); nl.setSpacing(0)
        self.buttons = {}
        for key, label in [("setup", "⚙   Pengaturan"), ("deploy", "🚀   Deploy"),
                           ("dashboard", "▦   Dasbor"), ("monitor", "📈   Pemantau"),
                           ("logs", "📋   Catatan")]:
            b = QPushButton(label); b.setProperty("active", "false")
            self.buttons[key] = b; nl.addWidget(b)
        lay.addWidget(nav); lay.addWidget(Divider())

        # ── Action color legend — fills the idle sidebar space with something
        # useful instead of a bare void, and ties the color language together ──
        legend = QWidget(); legend.setObjectName("sidebar_legend")
        ll = QVBoxLayout(legend); ll.setContentsMargins(18, 14, 18, 14); ll.setSpacing(9)
        lt = QLabel("LEGENDA AKSI"); lt.setObjectName("legend_title")
        ll.addWidget(lt)
        for key, label in [("maintain", "Maintain"), ("block_ip", "Block IP"),
                           ("tighten", "Tighten"), ("rollback", "Rollback")]:
            row = QWidget(); rlay = QHBoxLayout(row)
            rlay.setContentsMargins(0, 0, 0, 0); rlay.setSpacing(9)
            dot = QLabel(); dot.setFixedSize(9, 9)
            dot.setStyleSheet(f"background-color:{ACTION_COLORS[key]}; border-radius:4px;")
            txt = QLabel(label); txt.setObjectName("legend_item")
            rlay.addWidget(dot); rlay.addWidget(txt); rlay.addStretch()
            ll.addWidget(row)
        lay.addWidget(legend)

        lay.addStretch(); lay.addWidget(Divider())

        self._conn = QLabel("⚫  Belum terhubung")
        self._conn.setStyleSheet(f"color:{FAINT}; font-size:11px; padding:12px 18px;")
        lay.addWidget(self._conn)

    def set_active(self, key):
        for k, b in self.buttons.items():
            b.setProperty("active", "true" if k == key else "false")
            b.style().unpolish(b); b.style().polish(b)

    def set_connected(self, ok, text=None):
        if ok:
            self._conn.setText(text or "🟢  OPNsense terhubung")
            self._conn.setStyleSheet(f"color:{GREEN}; font-size:11px; padding:12px 18px;")
        else:
            self._conn.setText(text or "🔴  OPNsense offline")
            self._conn.setStyleSheet(f"color:{RED}; font-size:11px; padding:12px 18px;")


# ────────────────────────────────────────────────────────────────────────────
# Setup page
# ────────────────────────────────────────────────────────────────────────────

class SetupPage(QWidget):
    saved = pyqtSignal()

    def __init__(self, cfg: ConfigManager):
        super().__init__()
        self.cfg = cfg
        self._runner = None

        outer = QVBoxLayout(self); outer.setContentsMargins(0, 0, 0, 0)
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        outer.addWidget(scroll)
        body = QWidget(); scroll.setWidget(body)
        root = QVBoxLayout(body); root.setContentsMargins(24, 22, 24, 22); root.setSpacing(16)

        intro = QLabel("Konfigurasikan koneksi ke OPNsense Anda. Nilai disimpan di "
                       "komputer ini dan dipakai untuk semua deployment.")
        intro.setStyleSheet(f"color:{MUTED};"); intro.setWordWrap(True)
        root.addWidget(intro)

        # ── Connection card ──
        conn = self._card("KONEKSI OPNSENSE")
        cg = QGridLayout(); cg.setHorizontalSpacing(16); cg.setVerticalSpacing(14)
        self.f_host   = QLineEdit(cfg.get("opnsense_host"))
        self.f_host.setPlaceholderText("https://192.168.1.1:8443")
        self.f_key    = QLineEdit(cfg.get("api_key"));    self.f_key.setPlaceholderText("API Key")
        self.f_secret = QLineEdit(cfg.get("api_secret")); self.f_secret.setPlaceholderText("API Secret")
        self.f_secret.setEchoMode(QLineEdit.EchoMode.Password)
        self.f_verify = QCheckBox("Verifikasi sertifikat SSL")
        self.f_verify.setChecked(bool(cfg.get("verify_ssl")))
        cg.addWidget(Field("Host OPNsense", self.f_host), 0, 0, 1, 2)
        cg.addWidget(Field("API Key", self.f_key), 1, 0)
        cg.addWidget(Field("API Secret", self.f_secret), 1, 1)
        cg.addWidget(self.f_verify, 2, 0, 1, 2)
        conn.layout().addLayout(cg)
        root.addWidget(conn)

        # ── Firewall objects card ──
        fw = self._card("OBJEK FIREWALL")
        fg = QGridLayout(); fg.setHorizontalSpacing(16); fg.setVerticalSpacing(14)
        self.f_wan   = QLineEdit(cfg.get("wan_ip"))
        self.f_wanif = QLineEdit(cfg.get("wan_if"))
        self.f_wanif.setPlaceholderText("wan")
        self.f_wanif.setToolTip("Nama interface logis OPNsense tempat rule dibuat (mis. wan).")
        self.f_wanphys = QLineEdit(cfg.get("wan_phys_if"))
        self.f_wanphys.setPlaceholderText("em0")
        self.f_wanphys.setToolTip("Nama NIC fisik seperti di log firewall, trafik non-WAN diabaikan.")
        self.f_alias = QLineEdit(cfg.get("blocklist_alias"))
        self.f_rin   = QLineEdit(cfg.get("rule_tighten_inbound"))
        self.f_rin.setPlaceholderText("UUID rule (opsional)")
        self.f_rout  = QLineEdit(cfg.get("rule_tighten_outbound"))
        self.f_rout.setPlaceholderText("UUID rule (opsional)")
        fg.addWidget(Field("WAN IP (self-traffic)", self.f_wan), 0, 0)
        fg.addWidget(Field("Alias Blocklist", self.f_alias), 0, 1)
        fg.addWidget(Field("Interface WAN (rule)", self.f_wanif), 1, 0)
        fg.addWidget(Field("Interface fisik WAN (log)", self.f_wanphys), 1, 1)
        fg.addWidget(Field("Rule Tighten Inbound", self.f_rin), 2, 0)
        fg.addWidget(Field("Rule Tighten Outbound", self.f_rout), 2, 1)
        fw.layout().addLayout(fg)
        root.addWidget(fw)

        # ── Model & timing card ──
        mt = self._card("MODEL & WAKTU")
        mg = QGridLayout(); mg.setHorizontalSpacing(16); mg.setVerticalSpacing(14)
        # Δt dikunci 30 dtk: fitur elapsed_time model dilatih pada window 30 detik
        # dan di-hardcode di collector — mengubahnya merusak distribusi input model.
        self.f_delta = QSpinBox(); self.f_delta.setRange(30, 30); self.f_delta.setValue(30)
        self.f_delta.setSuffix("  detik"); self.f_delta.setReadOnly(True)
        self.f_delta.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.f_delta.setToolTip("Δt dikunci 30 detik agar sesuai jendela pelatihan model FireRL.")
        self.f_model = QLineEdit(cfg.get("model_path"))
        self.f_model.setPlaceholderText("(kosong = model bawaan FireRL)")
        self.f_scaler = QLineEdit(cfg.get("scaler_path"))
        self.f_scaler.setPlaceholderText("(kosong = scaler bawaan FireRL)")
        mg.addWidget(Field("Interval Jendela (Δt)", self.f_delta), 0, 0)
        mg.addWidget(Field("Lokasi Model (opsional)", self.f_model), 0, 1)
        delta_note = QLabel("Terkunci 30 detik, jendela inferensi = jendela pelatihan model.")
        delta_note.setStyleSheet(f"color:{FAINT}; font-size:11px;"); delta_note.setWordWrap(True)
        mg.addWidget(delta_note, 1, 0)
        mg.addWidget(Field("Lokasi Scaler (opsional)", self.f_scaler), 1, 1)
        # Resolved model/scaler metadata (path actually used + size + mtime)
        self.lbl_modelinfo = QLabel(""); self.lbl_modelinfo.setObjectName("stage_desc")
        self.lbl_modelinfo.setStyleSheet(f"color:{FAINT}; font-size:11px;")
        self.lbl_modelinfo.setWordWrap(True)
        mg.addWidget(self.lbl_modelinfo, 2, 0, 1, 2)
        # Auto-unblock TTL (gaya fail2ban bantime) — 0 = mati / hanya manual
        self.f_ttl = QSpinBox(); self.f_ttl.setRange(0, 3600)
        self.f_ttl.setValue(int(cfg.get("unblock_ttl_s", 0) or 0))
        self.f_ttl.setSuffix("  detik"); self.f_ttl.setSpecialValueText("Mati (manual)")
        self.f_ttl.setToolTip("Auto-unblock tiap IP setelah N detik aman tanpa serangan "
                              "(0 = mati; un-block hanya manual/Panik).")
        mg.addWidget(Field("Auto-unblock TTL", self.f_ttl), 3, 0)
        ttl_note = QLabel("0 = mati (disarankan). Bila diisi, IP dibuka otomatis setelah aman "
                          "selama durasi ini (gaya fail2ban bantime).")
        ttl_note.setStyleSheet(f"color:{FAINT}; font-size:11px;"); ttl_note.setWordWrap(True)
        mg.addWidget(ttl_note, 3, 1)
        mt.layout().addLayout(mg)
        root.addWidget(mt)
        self._refresh_model_info()
        self.f_model.textChanged.connect(lambda _: self._refresh_model_info())
        self.f_scaler.textChanged.connect(lambda _: self._refresh_model_info())

        # ── Actions ──
        row = QHBoxLayout()
        self.btn_test = QPushButton("Uji Koneksi"); self.btn_test.setObjectName("btn_ghost")
        self.btn_save = QPushButton("Simpan Konfigurasi"); self.btn_save.setObjectName("btn_primary")
        self.lbl_result = QLabel(""); self.lbl_result.setStyleSheet(f"color:{MUTED};")
        row.addWidget(self.btn_test); row.addWidget(self.btn_save)
        row.addWidget(self.lbl_result); row.addStretch()
        root.addLayout(row)
        root.addStretch()

        self.btn_save.clicked.connect(self._save)
        self.btn_test.clicked.connect(self._test)

    def _card(self, title):
        c = QWidget(); c.setObjectName("card")
        l = QVBoxLayout(c); l.setContentsMargins(18, 16, 18, 18); l.setSpacing(14)
        t = QLabel(title); t.setObjectName("card_title"); l.addWidget(t)
        return c

    def _collect(self):
        self.cfg.set("opnsense_host", self.f_host.text().strip())
        self.cfg.set("api_key",       self.f_key.text().strip())
        self.cfg.set("api_secret",    self.f_secret.text().strip())
        self.cfg.set("verify_ssl",    self.f_verify.isChecked())
        self.cfg.set("wan_ip",        self.f_wan.text().strip())
        self.cfg.set("wan_if",        self.f_wanif.text().strip() or "wan")
        self.cfg.set("wan_phys_if",   self.f_wanphys.text().strip() or "em0")
        self.cfg.set("blocklist_alias", self.f_alias.text().strip())
        self.cfg.set("rule_tighten_inbound",  self.f_rin.text().strip())
        self.cfg.set("rule_tighten_outbound", self.f_rout.text().strip())
        self.cfg.set("delta_t",       30)   # dikunci 30 dtk sesuai window pelatihan model
        self.cfg.set("model_path",    self.f_model.text().strip())
        self.cfg.set("scaler_path",   self.f_scaler.text().strip())
        self.cfg.set("unblock_ttl_s", int(self.f_ttl.value()))

    def _save(self):
        self._collect(); self.cfg.save()
        self.lbl_result.setText("✔ Konfigurasi tersimpan.")
        self.lbl_result.setStyleSheet(f"color:{GREEN};")
        self.saved.emit()

    def _test(self):
        self._collect()
        self.lbl_result.setText("⏳ Menguji koneksi…")
        self.lbl_result.setStyleSheet(f"color:{AMBER};")
        self.btn_test.setEnabled(False)
        self._runner = StageRunner(self.cfg, ["connection"])
        self._runner.stage_result.connect(self._test_done)
        self._runner.start()

    def _test_done(self, key, status, msg):
        color = STATUS_COLORS.get(status, MUTED)
        icon = {"ok": "✔", "warn": "⚠", "error": "✖"}.get(status, "•")
        self.lbl_result.setText(f"{icon} {msg}")
        self.lbl_result.setStyleSheet(f"color:{color};")
        self.btn_test.setEnabled(True)

    def refresh_from_cfg(self):
        """Re-load every field from config (e.g. after auto-provisioning fills in
        the rule UUIDs, or the file was edited by isi_config.py). Full refresh so
        no field silently drifts from the saved config."""
        self.f_host.setText(self.cfg.get("opnsense_host") or "")
        self.f_key.setText(self.cfg.get("api_key") or "")
        self.f_secret.setText(self.cfg.get("api_secret") or "")
        self.f_verify.setChecked(bool(self.cfg.get("verify_ssl")))
        self.f_wan.setText(self.cfg.get("wan_ip") or "")
        self.f_wanif.setText(self.cfg.get("wan_if") or "")
        self.f_wanphys.setText(self.cfg.get("wan_phys_if") or "")
        self.f_alias.setText(self.cfg.get("blocklist_alias") or "")
        self.f_rin.setText(self.cfg.get("rule_tighten_inbound") or "")
        self.f_rout.setText(self.cfg.get("rule_tighten_outbound") or "")
        self.f_model.setText(self.cfg.get("model_path") or "")
        self.f_scaler.setText(self.cfg.get("scaler_path") or "")
        self.f_ttl.setValue(int(self.cfg.get("unblock_ttl_s", 0) or 0))
        self._refresh_model_info()

    def _refresh_model_info(self):
        """Show which model/scaler files will actually be loaded, with size + mtime,
        so the user can confirm the deployed artefacts at a glance."""
        # reflect the fields the user is currently editing before resolving
        self.cfg.set("model_path",  self.f_model.text().strip())
        self.cfg.set("scaler_path", self.f_scaler.text().strip())

        def describe(path):
            if not path:
                return "— (tidak ditemukan)"
            try:
                st = os.stat(path)
                mb = st.st_size / (1024 * 1024)
                mt = datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")
                return f"{path}  ·  {mb:.1f} MB  ·  {mt}"
            except OSError:
                return f"{path}  ·  ⚠ file tidak ada"

        model = self.cfg.model_path()
        scaler = self.cfg.scaler_path()
        self.lbl_modelinfo.setText(
            f"Model aktif:  {describe(model)}\nScaler aktif:  {describe(scaler)}")


# ────────────────────────────────────────────────────────────────────────────
# Deploy page — stage selection + run + start monitoring
# ────────────────────────────────────────────────────────────────────────────

class StageRow(QWidget):
    """A selectable deployment stage. The WHOLE row is clickable — clicking the
    name/description toggles the checkbox — and the checked state is highlighted."""
    def __init__(self, key, name, desc, checked):
        super().__init__()
        self.key = key
        self.setObjectName("stage_row")
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        lay = QHBoxLayout(self); lay.setContentsMargins(14, 12, 16, 12); lay.setSpacing(12)
        self.chk = QCheckBox(); self.chk.setChecked(checked)
        self.chk.setCursor(Qt.CursorShape.PointingHandCursor)
        txt = QWidget(); tl = QVBoxLayout(txt); tl.setContentsMargins(0, 0, 0, 0); tl.setSpacing(2)
        n = QLabel(name); n.setObjectName("stage_name")
        d = QLabel(desc); d.setObjectName("stage_desc")
        tl.addWidget(n); tl.addWidget(d)
        self.status = QLabel("—"); self.status.setObjectName("stage_status")
        self.status.setStyleSheet(f"color:{FAINT};")
        self.status.setMinimumWidth(300)
        self.status.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.status.setWordWrap(True)
        lay.addWidget(self.chk); lay.addWidget(txt, 1); lay.addWidget(self.status)

        self.chk.toggled.connect(self._sync_checked)
        self._sync_checked(self.chk.isChecked())

    def mousePressEvent(self, e):
        # click anywhere on the row (except directly on the checkbox) toggles it
        if e.button() == Qt.MouseButton.LeftButton \
                and not self.chk.geometry().contains(e.pos()):
            self.chk.toggle()
        super().mousePressEvent(e)

    def _sync_checked(self, on):
        self.setProperty("checked", "true" if on else "false")
        self.style().unpolish(self); self.style().polish(self)

    def set_status(self, status, msg):
        colors = {"ok": GREEN, "warn": AMBER, "error": RED, "running": ORANGE}
        icon = {"ok": "✔", "warn": "⚠", "error": "✖", "running": "⏳"}.get(status, "•")
        self.status.setText(f"{icon} {msg}")
        self.status.setStyleSheet(f"color:{colors.get(status, MUTED)};")


class DeployPage(QWidget):
    start_requested = pyqtSignal()
    log_line = pyqtSignal(str)
    provisioned = pyqtSignal(dict)     # emits created UUIDs so Setup can refresh
    ready_changed = pyqtSignal(bool)   # deployment prerequisites pass/fail → header toggle

    def __init__(self, cfg: ConfigManager):
        super().__init__()
        self.cfg = cfg
        self._runner = None
        self._prov = None
        self._model_ok = False
        self._can_start = False   # prasyarat deployment sudah lolos → Start boleh aktif

        outer = QVBoxLayout(self); outer.setContentsMargins(0, 0, 0, 0)
        scroll = QScrollArea(); scroll.setWidgetResizable(True); outer.addWidget(scroll)
        body = QWidget(); scroll.setWidget(body)
        root = QVBoxLayout(body); root.setContentsMargins(24, 22, 24, 22); root.setSpacing(16)

        intro = QLabel("Pilih tahap deployment yang ingin dijalankan, lalu validasi. "
                       "Setelah tahap wajib lolos, mulai pemantauan dengan model FireRL.")
        intro.setStyleSheet(f"color:{MUTED};"); intro.setWordWrap(True)
        root.addWidget(intro)

        # ── Auto-setup OPNsense (untuk user yang belum paham OPNsense) ──────────
        prov = QWidget(); prov.setObjectName("card")
        pl = QVBoxLayout(prov); pl.setContentsMargins(16, 16, 16, 16); pl.setSpacing(10)
        pt = QLabel("PENYIAPAN OTOMATIS OPNSENSE"); pt.setObjectName("card_title"); pl.addWidget(pt)
        pdesc = QLabel(
            "Belum menyiapkan firewall di OPNsense? Tombol ini membuat otomatis semua "
            "objek yang dibutuhkan DQN: alias <b>rl_blocklist</b>, <b>rule block</b> "
            "yang menegakkan blokir, alias port <b>rl_tighten_ports</b> (22/3389/445), "
            "serta <b>rule tighten</b> pada port sensitif tersebut (inbound &amp; outbound, "
            "awalnya nonaktif, bukan blokir semua trafik). Aman dijalankan berulang, "
            "objek yang sudah ada tidak diduplikasi.")
        pdesc.setStyleSheet(f"color:{MUTED};"); pdesc.setWordWrap(True); pl.addWidget(pdesc)
        prow = QHBoxLayout()
        self.btn_provision = QPushButton("🛠  Siapkan OPNsense Otomatis")
        self.btn_provision.setObjectName("btn_ghost")
        self.btn_provision.setCursor(Qt.CursorShape.PointingHandCursor)
        self.lbl_prov = QLabel(""); self.lbl_prov.setStyleSheet(f"color:{MUTED};")
        self.lbl_prov.setWordWrap(True)
        prow.addWidget(self.btn_provision); prow.addWidget(self.lbl_prov, 1)
        pl.addLayout(prow)
        root.addWidget(prov)

        card = QWidget(); card.setObjectName("card")
        cl = QVBoxLayout(card); cl.setContentsMargins(16, 16, 16, 16); cl.setSpacing(10)
        t = QLabel("TAHAP DEPLOYMENT"); t.setObjectName("card_title"); cl.addWidget(t)

        self.rows = {}
        enabled = cfg.get("stages_enabled")
        for key in STAGE_ORDER:
            name, desc = STAGE_META[key]
            r = StageRow(key, name, desc, enabled.get(key, True))
            self.rows[key] = r; cl.addWidget(r)
        root.addWidget(card)

        row = QHBoxLayout()
        self.btn_run = QPushButton("Jalankan Tahap Terpilih"); self.btn_run.setObjectName("btn_ghost")
        self.btn_run.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_start = QPushButton("▶  Mulai Pemantauan"); self.btn_start.setObjectName("btn_primary")
        self.btn_start.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_start.setEnabled(False)
        row.addWidget(self.btn_run); row.addStretch(); row.addWidget(self.btn_start)
        root.addLayout(row)
        root.addStretch()

        self.btn_run.clicked.connect(self._run_stages)
        self.btn_start.clicked.connect(self.start_requested.emit)
        self.btn_provision.clicked.connect(self._provision)

    # ── Auto-provision OPNsense objects ────────────────────────────────────────
    def _provision(self):
        if not self.cfg.is_configured():
            self.lbl_prov.setText("⚠ Isi & simpan koneksi di halaman Pengaturan dulu.")
            self.lbl_prov.setStyleSheet(f"color:{AMBER};")
            return
        from provision import ProvisionRunner
        self.btn_provision.setEnabled(False)
        self.lbl_prov.setText("⏳ Membuat objek di OPNsense…")
        self.lbl_prov.setStyleSheet(f"color:{AMBER};")
        self.log_line.emit("[Setup] Memulai penyiapan otomatis OPNsense…")
        self._prov = ProvisionRunner(self.cfg, wan_if=self.cfg.get("wan_if") or "wan")
        self._prov.log_line.connect(lambda m: self.log_line.emit(f"[Setup] {m}"))
        self._prov.done.connect(self._provision_done)
        self._prov.start()

    def _provision_done(self, res):
        self.btn_provision.setEnabled(True)
        if res.get("ok"):
            # persist any created rule UUIDs so 'tighten' works at runtime
            if res.get("tighten_inbound"):
                self.cfg.set("rule_tighten_inbound", res["tighten_inbound"])
            if res.get("tighten_outbound"):
                self.cfg.set("rule_tighten_outbound", res["tighten_outbound"])
            if res.get("alias"):
                self.cfg.set("blocklist_alias", res["alias"])
            self.cfg.save()
            self.lbl_prov.setText("✔ OPNsense siap, alias & aturan dibuat.")
            self.lbl_prov.setStyleSheet(f"color:{GREEN};")
            self.provisioned.emit(res)
        else:
            self.lbl_prov.setText("⚠ Sebagian gagal, cek Catatan untuk detail.")
            self.lbl_prov.setStyleSheet(f"color:{RED};")

    def _selected(self):
        return [k for k, r in self.rows.items() if r.chk.isChecked()]

    def _run_stages(self):
        sel = self._selected()
        if not sel:
            self.log_line.emit("[Deploy] Tidak ada tahap yang dipilih.")
            return
        # persist selection
        self.cfg.set("stages_enabled", {k: self.rows[k].chk.isChecked() for k in STAGE_ORDER})
        self.cfg.save()

        self.btn_run.setEnabled(False); self.btn_start.setEnabled(False)
        self._model_ok = False
        self._can_start = False
        self.ready_changed.emit(False)   # checking → header Start off until done
        for k in sel:
            self.rows[k].set_status("running", "menunggu…")
        self.log_line.emit(f"[Deploy] Menjalankan {len(sel)} tahap: {', '.join(sel)}")

        self._runner = StageRunner(self.cfg, sel)
        self._runner.stage_started.connect(lambda k: self.rows[k].set_status("running", "menjalankan…"))
        self._runner.stage_result.connect(self._on_result)
        self._runner.all_done.connect(self._on_all_done)
        self._runner.start()

    def _on_result(self, key, status, msg):
        self.rows[key].set_status(status, msg)
        self.log_line.emit(f"[Deploy] {key}: {msg}")
        if key == "model" and status == "ok":
            self._model_ok = True

    def _on_all_done(self, all_ok):
        self.btn_run.setEnabled(True)
        # Model must be loadable to run; connection must not have errored.
        conn_ok = self.rows["connection"].status.text().startswith("✔") \
            if self.rows["connection"].chk.isChecked() else True
        model_sel = self.rows["model"].chk.isChecked()
        can_start = (self._model_ok or not model_sel) and conn_ok
        self._can_start = can_start
        self.btn_start.setEnabled(can_start)
        self.ready_changed.emit(can_start)
        if can_start:
            self.log_line.emit("[Deploy] Prasyarat lolos, siap mulai pemantauan.")
        else:
            self.log_line.emit("[Deploy] Ada tahap wajib yang gagal, perbaiki dulu.")

    def mark_running(self, running):
        # Saat berjalan: Start dimatikan. Saat berhenti: pulihkan berdasarkan
        # apakah prasyarat terakhir lolos — tidak perlu cek tahap ulang.
        self.btn_start.setEnabled(self._can_start and not running)
        self.btn_run.setEnabled(not running)


# ────────────────────────────────────────────────────────────────────────────
# Dashboard page
# ────────────────────────────────────────────────────────────────────────────

class DashboardPage(QWidget):
    unblock_ip_requested  = pyqtSignal(str)   # minta buka blokir 1 IP (per-IP)
    unblock_all_requested = pyqtSignal()      # minta buka semua blokir (pemulihan)

    def __init__(self, lat_warn_ms=1200):
        super().__init__()
        self.lat_warn_ms = int(lat_warn_ms)
        root = QVBoxLayout(self); root.setContentsMargins(20, 18, 20, 18); root.setSpacing(14)

        # Banner rekomendasi un-block (pola SOAR one-click rollback) — muncul saat aman
        self.banner = QFrame(); self.banner.setObjectName("banner_unblock")
        self.banner.setStyleSheet(
            "#banner_unblock{background:#3a2410;border:1px solid #e07b39;border-radius:8px;}")
        _bl = QHBoxLayout(self.banner); _bl.setContentsMargins(14, 10, 14, 10); _bl.setSpacing(12)
        self.banner_lbl = QLabel(""); self.banner_lbl.setStyleSheet("color:#ffb27a;font-weight:600;")
        self.banner_btn = QPushButton("Buka Blokir"); self.banner_btn.setObjectName("btn_primary")
        self.banner_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.banner_btn.clicked.connect(self.unblock_all_requested.emit)
        _bl.addWidget(self.banner_lbl, 1); _bl.addWidget(self.banner_btn, 0)
        self.banner.setVisible(False)
        root.addWidget(self.banner)

        cards = QHBoxLayout(); cards.setSpacing(12)
        self.c_win   = StatCard("Jendela", "0", BLUE, icon="▦")
        self.c_act   = StatCard("Aksi Terakhir", "—", TEXT, icon="⚡")
        self.c_lat   = StatCard("Latensi", "—", AMBER, icon="⏱")
        self.c_blk   = StatCard("Blocklist", "0", ORANGE, icon="⛔")
        self.c_guard = StatCard("Guard Aktif", "0", "#a56cd6", icon="🛡")
        for c in (self.c_win, self.c_act, self.c_lat, self.c_blk, self.c_guard):
            cards.addWidget(c)
        root.addLayout(cards)

        charts = QHBoxLayout(); charts.setSpacing(12)
        qw = self._wrap("Q-VALUES"); self.q_chart = QValueChart(); qw.layout().addWidget(self.q_chart)
        charts.addWidget(qw, 3)
        aw = self._wrap("DISTRIBUSI AKSI"); self.a_chart = ActionDistChart(); aw.layout().addWidget(self.a_chart)
        charts.addWidget(aw, 2)
        root.addLayout(charts)

        tw = self._wrap("JENDELA TERBARU")
        self.table = QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels(
            ["Jendela", "Waktu", "Aksi", "Latensi (ms)", "Blocklist", "Q-Maks", "Guard"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self._stack = _table_with_empty(self.table,
            "Belum ada keputusan, mulai pemantauan untuk melihat data jendela.")
        self._stack.setMinimumHeight(150); self._stack.setMaximumHeight(240)
        tw.layout().addWidget(self._stack)

        # Panel blocklist per-IP (untuk skenario banyak serangan) — unblock granular
        bw = self._wrap("IP DIBLOKIR")
        self.blk_table = QTableWidget(0, 3)
        self.blk_table.setHorizontalHeaderLabels(["IP", "Lama Diblokir", "Aksi"])
        _bh = self.blk_table.horizontalHeader()
        _bh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)   # IP
        _bh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)   # Lama Diblokir
        # Aksi: lebar TETAP (ResizeToContents tak menghitung cell-widget → tombol meluber).
        _bh.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.blk_table.setColumnWidth(2, 190)
        self.blk_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.blk_table.verticalHeader().setVisible(False)
        self.blk_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.blk_table.verticalHeader().setDefaultSectionSize(46)   # ruang cukup utk tombol (anti-terpotong bawah)
        self.blk_stack = _table_with_empty(self.blk_table, "Tidak ada IP diblokir.")
        self.blk_stack.setMinimumHeight(150); self.blk_stack.setMaximumHeight(240)
        bw.layout().addWidget(self.blk_stack)

        # Tata letak berdampingan: kiri JENDELA TERBARU (lebih lebar, 7 kolom),
        # kanan IP DIBLOKIR (3 kolom). Tiap tabel scroll internal saat baris banyak.
        bottom = QHBoxLayout(); bottom.setSpacing(12)
        bottom.addWidget(tw, 3)
        bottom.addWidget(bw, 2)
        root.addLayout(bottom)

        # "Lama Diblokir" berdetak per-detik (bukan hanya tiap window 30 dtk).
        self._blk_start = {}   # ip -> epoch mulai diblokir
        self._dur_timer = QTimer(self); self._dur_timer.setInterval(1000)
        self._dur_timer.timeout.connect(self._tick_durations)
        self._dur_timer.start()

    def _wrap(self, title):
        w = QWidget(); w.setObjectName("card")
        l = QVBoxLayout(w); l.setContentsMargins(12, 12, 12, 12); l.setSpacing(8)
        t = QLabel(title); t.setObjectName("card_title"); l.addWidget(t)
        return w

    def reset(self):
        self.q_chart.reset(); self.a_chart.reset()
        self.table.setRowCount(0); self._stack.setCurrentIndex(0)
        self.banner.setVisible(False)
        self.blk_table.setRowCount(0); self.blk_stack.setCurrentIndex(0); self._blk_start = {}
        for c, v in ((self.c_win, "0"), (self.c_act, "—"), (self.c_lat, "—"),
                     (self.c_blk, "0"), (self.c_guard, "0")):
            c.set_value(v)

    def update(self, wr):
        color = ACTION_COLORS.get(wr.action_name, "#fff")
        self._stack.setCurrentIndex(1)
        self.c_win.set_value(str(wr.window_id + 1), BLUE)
        self.c_act.set_value(wr.action_name, color)
        self.c_lat.set_value(f"{wr.latency_ms:.0f} ms",
                             RED if wr.latency_ms > self.lat_warn_ms else AMBER)
        self.c_blk.set_value(str(wr.blocklist_size), ORANGE if wr.blocklist_size else GREEN)
        self.q_chart.push(wr.window_id, wr.q_values)
        self.a_chart.increment(wr.action)

        r = self.table.rowCount(); self.table.insertRow(r)
        qmax = max(wr.q_values) if wr.q_values else 0
        cells = [str(wr.window_id), wr.timestamp, wr.action_name,
                 f"{wr.latency_ms:.1f}", str(wr.blocklist_size),
                 f"{qmax:.2f}", "🛡" if wr.guard_fired else "—"]
        for col, txt in enumerate(cells):
            it = QTableWidgetItem(txt); it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if col == 2: it.setForeground(QColor(color))
            if col == 6 and wr.guard_fired: it.setForeground(QColor("#a56cd6"))
            self.table.setItem(r, col, it)
        self.table.scrollToBottom()

    def set_guard_count(self, n):
        self.c_guard.set_value(str(n), "#a56cd6" if n else FAINT)

    def set_unblock_hint(self, ready, blk, safe):
        """Tampilkan/sembunyikan banner rekomendasi un-block (dari worker.rollback_hint)."""
        if ready and blk > 0:
            self.banner_lbl.setText(
                f"✅ Aman membuka blokir — {blk} IP diblokir, {safe} jendela aman "
                f"berturut-turut tanpa serangan.")
            self.banner.setVisible(True)
        else:
            self.banner.setVisible(False)

    @staticmethod
    def _fmt_dur(age):
        age = max(0, int(age)); m, s = divmod(age, 60)
        return f"{m}m {s}s" if m else f"{s}s"

    def _tick_durations(self):
        """Dipanggil QTimer tiap 1 detik — perbarui kolom 'Lama Diblokir' agar
        berdetak hidup, tanpa membangun ulang tabel (tombol tetap utuh)."""
        if self.blk_table.rowCount() == 0:
            return
        now = time.time()
        for r in range(self.blk_table.rowCount()):
            ipit = self.blk_table.item(r, 0)
            start = self._blk_start.get(ipit.text()) if ipit else None
            cell = self.blk_table.item(r, 1)
            if start and cell:
                cell.setText(self._fmt_dur(now - start))

    def update_blocklist(self, rows):
        """Isi tabel IP diblokir (dari worker.blocklist_changed). rows=[(ip, epoch_mulai),…].
        Durasi dihitung dari epoch agar bisa berdetak per-detik via _tick_durations."""
        new_start = {ip: since for ip, since in rows}
        # Himpunan IP tak berubah → cukup segarkan waktu-mulai, biar timer yang
        # mendetakkan durasi. Hindari bangun ulang tabel + tombol tiap window.
        if set(new_start) == set(self._blk_start):
            self._blk_start = new_start
            return
        self._blk_start = new_start
        self.blk_table.setRowCount(0)
        if not rows:
            self.blk_stack.setCurrentIndex(0)
            return
        self.blk_stack.setCurrentIndex(1)
        now = time.time()
        for ip, since in rows:
            r = self.blk_table.rowCount(); self.blk_table.insertRow(r)
            dur = self._fmt_dur((now - since) if since else 0)
            for col, txt in ((0, ip), (1, dur)):
                it = QTableWidgetItem(txt); it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.blk_table.setItem(r, col, it)
            # Tombol kompak dgn style sendiri (BUKAN btn_ghost yg padding 10px-nya
            # membuat teks terpotong di baris pendek → tombol tampak kotak kosong).
            cell = QWidget()
            cl = QHBoxLayout(cell); cl.setContentsMargins(8, 4, 12, 4); cl.setSpacing(0)
            btn = QPushButton("🔓  Buka Blokir")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setToolTip(f"Buka blokir {ip} — hapus dari blocklist OPNsense.")
            btn.setMinimumWidth(140)   # cukup utk teks penuh (anti-terpotong horizontal)
            btn.setStyleSheet(
                f"QPushButton{{color:{ORANGE}; background:transparent;"
                f"border:1px solid {ORANGE}; border-radius:6px; padding:5px 16px;"
                f"font-size:12px; font-weight:600;}}"
                f"QPushButton:hover{{background-color:rgba(232,98,42,0.16);}}")
            btn.clicked.connect(lambda _, a=ip: self.unblock_ip_requested.emit(a))
            cl.addWidget(btn, 0, Qt.AlignmentFlag.AlignCenter)
            self.blk_table.setCellWidget(r, 2, cell)

    def remove_blocklist_row(self, ip):
        """Hapus baris IP seketika (optimistic) begitu 'Buka' diklik — tak menunggu
        window berikutnya. Sumber kebenaran tetap worker.blocklist_changed."""
        self._blk_start.pop(ip, None)   # jaga konsistensi dgn optimasi set-IP
        for r in range(self.blk_table.rowCount()):
            it = self.blk_table.item(r, 0)
            if it and it.text() == ip:
                self.blk_table.removeRow(r)
                break
        if self.blk_table.rowCount() == 0:
            self.blk_stack.setCurrentIndex(0)


# ────────────────────────────────────────────────────────────────────────────
# Monitor page
# ────────────────────────────────────────────────────────────────────────────

class MonitorPage(QWidget):
    def __init__(self):
        super().__init__()
        root = QVBoxLayout(self); root.setContentsMargins(20, 18, 20, 18); root.setSpacing(12)

        qw = QWidget(); qw.setObjectName("card")
        ql = QVBoxLayout(qw); ql.setContentsMargins(10, 10, 10, 10); ql.setSpacing(6)
        t = QLabel("Q-VALUE · LANGSUNG"); t.setObjectName("card_title"); ql.addWidget(t)
        self.q_chart = QValueChart(height=300); ql.addWidget(self.q_chart)
        root.addWidget(qw, 2)

        info = QWidget(); info.setObjectName("card")
        il = QHBoxLayout(info); il.setContentsMargins(16, 11, 16, 11); il.setSpacing(14)
        self.lbl_q = QLabel("Q: —"); self.lbl_atk = QLabel("Serangan: —")
        self.lbl_safe = QLabel("Aman: 0"); self.lbl_guard = QLabel("Guard: siaga")
        strip = [self.lbl_q, self.lbl_atk, self.lbl_safe, self.lbl_guard]
        for i, w in enumerate(strip):
            w.setStyleSheet(f"color:{MUTED}; font-size:12px;"); il.addWidget(w)
            if i < len(strip) - 1:
                il.addWidget(Divider(vertical=True))
        il.addStretch()
        root.addWidget(info)

        tw = QWidget(); tw.setObjectName("card")
        tl = QVBoxLayout(tw); tl.setContentsMargins(12, 12, 12, 12); tl.setSpacing(8)
        t2 = QLabel("SEMUA KEPUTUSAN"); t2.setObjectName("card_title"); tl.addWidget(t2)
        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels(
            ["Jdl", "Waktu", "Aksi", "Lat(ms)", "Blk", "Q₀", "Q₁", "Q₂/Q₃"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self._stack = _table_with_empty(self.table,
            "Menunggu keputusan pertama, grafik & tabel akan terisi tiap jendela (Δt).")
        tl.addWidget(self._stack)
        root.addWidget(tw, 3)

    def reset(self):
        self.q_chart.reset(); self.table.setRowCount(0); self._stack.setCurrentIndex(0)

    def update(self, wr, consec_safe):
        self._stack.setCurrentIndex(1)
        color = ACTION_COLORS.get(wr.action_name, "#fff"); q = wr.q_values
        self.q_chart.push(wr.window_id, q)
        if q: self.lbl_q.setText("  ".join(f"Q{i}={v:.2f}" for i, v in enumerate(q)))
        self.lbl_atk.setText("⚠ Serangan" if wr.is_attack else "✔ Aman")
        self.lbl_atk.setStyleSheet(f"color:{RED if wr.is_attack else GREEN};")
        self.lbl_safe.setText(f"Aman: {consec_safe}")
        if wr.guard_fired:
            self.lbl_guard.setText("🛡 Guard AKTIF"); self.lbl_guard.setStyleSheet("color:#a56cd6; font-weight:bold;")
        else:
            self.lbl_guard.setText("Guard: siaga"); self.lbl_guard.setStyleSheet(f"color:{FAINT};")

        r = self.table.rowCount(); self.table.insertRow(r)
        q23 = f"{q[2]:.2f}/{q[3]:.2f}" if len(q) >= 4 else "—"
        cells = [str(wr.window_id), wr.timestamp, wr.action_name, f"{wr.latency_ms:.1f}",
                 str(wr.blocklist_size), f"{q[0]:.2f}" if q else "—",
                 f"{q[1]:.2f}" if len(q) > 1 else "—", q23]
        for col, txt in enumerate(cells):
            it = QTableWidgetItem(txt); it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if col == 2:
                it.setForeground(QColor(color))
                if wr.guard_fired:
                    it.setText("maintain 🛡"); it.setForeground(QColor("#9b59b6"))
            self.table.setItem(r, col, it)
        self.table.scrollToBottom()


# ────────────────────────────────────────────────────────────────────────────
# Logs page
# ────────────────────────────────────────────────────────────────────────────

class LogsPage(QWidget):
    def __init__(self):
        super().__init__()
        root = QVBoxLayout(self); root.setContentsMargins(20, 18, 20, 18); root.setSpacing(8)
        bar = QHBoxLayout()
        t = QLabel("CATATAN SISTEM"); t.setObjectName("card_title"); bar.addWidget(t); bar.addStretch()
        self.btn_clear = QPushButton("Bersihkan"); self.btn_clear.setObjectName("btn_clear")
        bar.addWidget(self.btn_clear); root.addLayout(bar)
        self.log = QPlainTextEdit(); self.log.setObjectName("log_panel")
        self.log.setReadOnly(True); self.log.setMaximumBlockCount(3000)
        root.addWidget(self.log)
        self.btn_clear.clicked.connect(self.log.clear)

    def append(self, msg):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        self.log.appendPlainText(f"[{ts}] {msg}")


# ────────────────────────────────────────────────────────────────────────────
# Main window
# ────────────────────────────────────────────────────────────────────────────

class FireSenseWindow(QMainWindow):
    def __init__(self, logo_path):
        super().__init__()
        self.setWindowTitle("FireSense · Pemantau Firewall DQN")
        self.setMinimumSize(1140, 740); self.resize(1320, 820)
        self.setWindowIcon(QIcon(logo_path))
        self.setStyleSheet(DARK_THEME)

        self.cfg = ConfigManager()
        self._guard = 0; self._safe = 0; self._worker = None
        self._ready = False   # deployment prerequisites currently satisfied?
        self._build(logo_path)

        # jump to Setup first if not configured yet
        self._switch("deploy" if self.cfg.is_configured() else "setup")

    def _build(self, logo_path):
        central = QWidget(); self.setCentralWidget(central)
        root = QHBoxLayout(central); root.setContentsMargins(0, 0, 0, 0); root.setSpacing(0)

        self.sidebar = Sidebar(logo_path); root.addWidget(self.sidebar)

        right = QWidget(); rl = QVBoxLayout(right); rl.setContentsMargins(0, 0, 0, 0); rl.setSpacing(0)

        hbar = QWidget(); hbar.setObjectName("header_bar")
        hl = QHBoxLayout(hbar); hl.setContentsMargins(22, 8, 22, 8); hl.setSpacing(12)
        titlebox = QWidget(); tb = QVBoxLayout(titlebox); tb.setContentsMargins(0, 0, 0, 0); tb.setSpacing(0)
        self.page_title = QLabel("Pengaturan"); self.page_title.setObjectName("section_title")
        self.page_sub = ElidedLabel(""); self.page_sub.setObjectName("section_sub")
        tb.addWidget(self.page_title); tb.addWidget(self.page_sub)
        hl.addWidget(titlebox, 1, Qt.AlignmentFlag.AlignVCenter)
        self.status_pill = QLabel("⚫  BERHENTI"); self.status_pill.setObjectName("status_stopped")
        hl.addWidget(self.status_pill, 0, Qt.AlignmentFlag.AlignVCenter)
        self.btn_export = QPushButton("⬇  Hasil Pemantauan"); self.btn_export.setObjectName("btn_ghost")
        self.btn_export.setToolTip("Lihat / ekspor hasil pemantauan (window_decisions.csv).")
        self.btn_export.setCursor(Qt.CursorShape.PointingHandCursor)
        hl.addWidget(self.btn_export, 0, Qt.AlignmentFlag.AlignVCenter)
        self.btn_unblock = QPushButton("🔓  Buka Blokir"); self.btn_unblock.setObjectName("btn_ghost")
        self.btn_unblock.setToolTip("Pemulihan normal: buka semua blokir & nonaktifkan tighten.")
        self.btn_unblock.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_unblock.setEnabled(False)
        hl.addWidget(self.btn_unblock, 0, Qt.AlignmentFlag.AlignVCenter)
        self.btn_panic = QPushButton("🛑  Panik"); self.btn_panic.setObjectName("btn_stop")
        self.btn_panic.setToolTip("DARURAT: buka SEMUA blokir & nonaktifkan rule tighten segera.")
        self.btn_panic.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_panic.setEnabled(False)
        hl.addWidget(self.btn_panic, 0, Qt.AlignmentFlag.AlignVCenter)
        # One button that toggles between Start (idle) and Stop (running).
        self.btn_toggle = QPushButton("▶  Mulai"); self.btn_toggle.setObjectName("btn_primary")
        self.btn_toggle.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_toggle.setEnabled(False)
        hl.addWidget(self.btn_toggle, 0, Qt.AlignmentFlag.AlignVCenter)
        rl.addWidget(hbar); rl.addWidget(Divider())

        self.stack = QStackedWidget()
        self.p_setup = SetupPage(self.cfg)
        self.p_deploy = DeployPage(self.cfg)
        self.p_dash = DashboardPage(self.cfg.get("latency_warn_ms", 1200))
        self.p_monitor = MonitorPage()
        self.p_logs = LogsPage()
        for p in (self.p_setup, self.p_deploy, self.p_dash, self.p_monitor, self.p_logs):
            self.stack.addWidget(p)
        rl.addWidget(self.stack)
        root.addWidget(right, 1)

        for k in ("setup", "deploy", "dashboard", "monitor", "logs"):
            self.sidebar.buttons[k].clicked.connect(lambda _, key=k: self._switch(key))
        self.btn_toggle.clicked.connect(self._toggle_run)
        self.btn_export.clicked.connect(self._export_results)
        self.btn_panic.clicked.connect(self._panic)
        self.btn_unblock.clicked.connect(self._unblock_all)
        self.p_dash.unblock_all_requested.connect(self._unblock_all)
        self.p_dash.unblock_ip_requested.connect(self._unblock_ip)
        self.p_deploy.start_requested.connect(self._start)
        self.p_deploy.ready_changed.connect(self._on_ready_changed)
        self.p_deploy.log_line.connect(self.p_logs.append)
        self.p_deploy.provisioned.connect(lambda res: self.p_setup.refresh_from_cfg())
        self.p_setup.saved.connect(lambda: self.p_logs.append("[Setup] Konfigurasi disimpan."))

    _TITLES = {
        "setup":     ("Pengaturan", "Konfigurasi koneksi OPNsense Anda"),
        "deploy":    ("Deploy", "Pilih & jalankan tahap deployment"),
        "dashboard": ("Dasbor", "Ringkasan keputusan waktu nyata"),
        "monitor":   ("Pemantau Langsung", "Q-value & keputusan per jendela"),
        "logs":      ("Catatan Sistem", "Catatan aktivitas aplikasi"),
    }
    _INDEX = {"setup": 0, "deploy": 1, "dashboard": 2, "monitor": 3, "logs": 4}

    def _switch(self, key):
        self.stack.setCurrentIndex(self._INDEX[key])
        self.sidebar.set_active(key)
        title, sub = self._TITLES[key]
        self.page_title.setText(title); self.page_sub.setText(sub)

    # ── deployment control ──
    def _start(self):
        # Confirm before touching a live firewall — this drives real OPNsense rules.
        ok = QMessageBox.question(
            self, "Mulai pemantauan?",
            "FireSense akan mulai menjalankan model DQN dan dapat mengubah rule "
            "firewall OPNsense Anda secara langsung (block IP / tighten port).\n\n"
            "Lanjutkan?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if ok != QMessageBox.StandardButton.Yes:
            self.p_logs.append("[FireSense] Pemantauan dibatalkan oleh pengguna.")
            return

        from worker import DeploymentWorker
        self._guard = 0; self._safe = 0
        self.p_dash.reset(); self.p_monitor.reset()
        self._set_toggle(True)
        self.btn_panic.setEnabled(True)
        self.btn_unblock.setEnabled(True)
        self.p_deploy.mark_running(True)
        self._set_status("connecting", "🟡  MENGHUBUNGKAN")
        self._switch("dashboard")

        self._worker = DeploymentWorker(self.cfg)
        self._worker.window_done.connect(self._on_window)
        self._worker.log_line.connect(self.p_logs.append)
        self._worker.connected.connect(self._on_connected)
        self._worker.finished_clean.connect(self._on_stopped)
        self._worker.error.connect(self._on_error)
        self._worker.rollback_hint.connect(self.p_dash.set_unblock_hint)
        self._worker.blocklist_changed.connect(self.p_dash.update_blocklist)
        self._worker.start()
        self.p_logs.append("[FireSense] Deployment dimulai.")

    def _toggle_run(self):
        """Header button: Stop the worker if running, otherwise Start it."""
        if self._worker and self._worker.isRunning():
            self._stop()
        else:
            self._start()

    def _on_ready_changed(self, ready):
        """Deploy page reports whether prerequisites pass; reflect on the header
        Start button — but never override while a run is in progress."""
        self._ready = ready
        if not (self._worker and self._worker.isRunning()):
            self._set_toggle(False)

    def _set_toggle(self, running):
        """Switch the single header button between Start and Stop states."""
        if running:
            self.btn_toggle.setText("■  Berhenti")
            self.btn_toggle.setObjectName("btn_stop")
            self.btn_toggle.setEnabled(True)
        else:
            self.btn_toggle.setText("▶  Mulai")
            self.btn_toggle.setObjectName("btn_primary")
            self.btn_toggle.setEnabled(self._ready)
        # re-apply stylesheet after objectName change
        self.btn_toggle.style().unpolish(self.btn_toggle)
        self.btn_toggle.style().polish(self.btn_toggle)

    def _stop(self):
        if self._worker: self._worker.stop()
        self.btn_toggle.setEnabled(False)   # disabled until the loop confirms stop
        self.p_logs.append("[FireSense] Permintaan berhenti…")

    def _panic(self):
        if not (self._worker and self._worker.isRunning()):
            return
        ok = QMessageBox.question(
            self, "Panik: buka semua blokir?",
            "Tindakan darurat: SEMUA IP di blocklist akan dibuka dan rule tighten "
            "dinonaktifkan, lalu perubahan diterapkan ke OPNsense.\n\n"
            "Deployment tetap berjalan. Lanjutkan?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if ok != QMessageBox.StandardButton.Yes:
            return
        n = len(self.p_dash._blk_start)     # membuka blokir = rollback (dieksekusi operator)
        if n: self.p_dash.a_chart.increment(3, n)
        self._worker.request_panic()
        self.p_logs.append(f"[FireSense] 🛑 Panik diminta, membuka {n} blokir (rollback)…")

    def _unblock_all(self):
        """Pemulihan normal (bukan darurat) — buka semua blokir + nonaktifkan tighten."""
        if not (self._worker and self._worker.isRunning()):
            return
        ok = QMessageBox.question(
            self, "Buka semua blokir?",
            "Semua IP di blocklist akan dibuka dan rule tighten dinonaktifkan "
            "(pemulihan ke kondisi normal). Deployment tetap berjalan.\n\nLanjutkan?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if ok != QMessageBox.StandardButton.Yes:
            return
        n = len(self.p_dash._blk_start)     # tiap IP dibuka dihitung 1 rollback
        if n: self.p_dash.a_chart.increment(3, n)
        self._worker.request_unblock_all()
        self.p_dash.banner.setVisible(False)
        self.p_dash.update_blocklist([])       # kosongkan tabel seketika (optimistic)
        self.p_logs.append(f"[FireSense] 🔓 Buka {n} blokir (rollback, pemulihan)…")

    def _unblock_ip(self, ip):
        """Buka blokir satu IP dari panel blocklist (rollback per-IP)."""
        if not (self._worker and self._worker.isRunning()):
            return
        self._worker.request_unblock_ip(ip)
        self.p_dash.a_chart.increment(3, 1)    # 1 IP dibuka = 1 rollback
        self.p_dash.remove_blocklist_row(ip)   # umpan balik seketika (optimistic)
        self.p_logs.append(f"[FireSense] 🔓 Buka blokir IP {ip} (rollback)…")

    def _on_window(self, wr):
        if wr.guard_fired: self._guard += 1; self._safe = 0
        elif not wr.is_attack: self._safe += 1
        else: self._safe = 0
        self.p_dash.update(wr); self.p_dash.set_guard_count(self._guard)
        self.p_monitor.update(wr, self._safe)

    def _on_connected(self, ok):
        self.sidebar.set_connected(ok)
        if ok: self._set_status("running", "🟢  BERJALAN")
        else:  self._set_status("stopped", "🔴  TAK TERHUBUNG")

    def _on_stopped(self):
        self._set_toggle(False)   # back to Start, enabled iff prerequisites still hold
        self.btn_panic.setEnabled(False)
        self.btn_unblock.setEnabled(False)
        self.p_dash.banner.setVisible(False)
        self.p_deploy.mark_running(False)
        self._set_status("stopped", "⚫  BERHENTI")
        self.sidebar.set_connected(False)
        self.p_logs.append("[FireSense] Deployment dihentikan.")

    def _on_error(self, msg):
        self.p_logs.append(f"[ERROR] {msg}")
        self._on_stopped()

    def _set_status(self, kind, text):
        obj = {"running": "status_running", "connecting": "status_connecting",
               "stopped": "status_stopped"}[kind]
        self.status_pill.setText(text); self.status_pill.setObjectName(obj)
        self.status_pill.style().unpolish(self.status_pill)
        self.status_pill.style().polish(self.status_pill)

    # ── monitoring results (cek / export) ──
    def _results_dir(self):
        """Folder where the worker writes window_decisions.csv — mirrors
        worker.DeploymentWorker._results_dir so export finds the same file."""
        if getattr(sys, "frozen", False):
            return os.path.join(os.path.dirname(sys.executable), "results")
        try:
            import opnsense_config as ocfg
            return os.path.dirname(ocfg.LOG_PATH)
        except Exception:
            return os.path.join(os.path.dirname(os.path.dirname(
                os.path.abspath(__file__))), "results")

    @staticmethod
    def _to_num(s):
        """Convert a CSV string cell to int/float when possible (for real Excel
        numeric cells), else return the original string."""
        if s is None or s == "":
            return s
        try:
            f = float(s)
        except (TypeError, ValueError):
            return s
        if f.is_integer() and "." not in str(s) and "e" not in str(s).lower():
            return int(f)
        return f

    def _build_results_xlsx(self, src_csv, dict_csv, dest_xlsx):
        """Build ONE styled Office Open XML (ISO/IEC 29500) workbook with three
        sheets: Ringkasan, Keputusan, Kamus Data. Raises on failure so the caller
        can fall back to plain CSV."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ORANGE, WHITE, ALT = "E8622A", "FFFFFF", "FCEDE6"
        hdr_font   = Font(bold=True, color=WHITE, size=11)
        hdr_fill   = PatternFill("solid", fgColor=ORANGE)
        title_font = Font(bold=True, size=16, color=ORANGE)
        lbl_font   = Font(bold=True, color="333333")
        alt_fill   = PatternFill("solid", fgColor=ALT)
        thin       = Side(style="thin", color="E2E2E2")
        border     = Border(left=thin, right=thin, top=thin, bottom=thin)
        center     = Alignment(horizontal="center", vertical="center")
        left_wrap  = Alignment(horizontal="left", vertical="center", wrap_text=True)

        def read_csv(path):
            with open(path, newline="", encoding="utf-8") as f:
                return list(csv.reader(f))

        def autosize(ws, header, rows, cap=60):
            for j in range(len(header)):
                width = len(str(header[j]))
                for r in rows:
                    if j < len(r):
                        width = max(width, len(str(r[j])))
                ws.column_dimensions[get_column_letter(j + 1)].width = min(width + 3, cap)

        def write_table(ws, header, rows, wrap_last=False):
            for j, h in enumerate(header, 1):
                c = ws.cell(1, j, h)
                c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, center, border
            for i, row in enumerate(rows, 2):
                for j, raw in enumerate(row, 1):
                    val = self._to_num(raw)
                    c = ws.cell(i, j, val)
                    c.border = border
                    is_num = isinstance(val, (int, float))
                    if wrap_last and j == len(header):
                        c.alignment = left_wrap
                    else:
                        c.alignment = center if is_num else Alignment(
                            horizontal="left", vertical="center")
                    if i % 2 == 1:
                        c.fill = alt_fill
            ws.freeze_panes = "A2"
            autosize(ws, header, rows)

        dec = read_csv(src_csv)
        d_header, d_rows = (dec[0], dec[1:]) if dec else ([], [])
        idx = {name: i for i, name in enumerate(d_header)}

        wb = Workbook()

        # ── Sheet 1: Ringkasan (mini dashboard) ───────────────────────────────
        from openpyxl.chart import BarChart, Reference
        from openpyxl.formatting.rule import DataBarRule

        ws = wb.active
        ws.title = "Ringkasan"
        ws.sheet_view.showGridLines = False

        banner_font  = Font(bold=True, color=WHITE, size=18)
        sub_font     = Font(italic=True, color="8A8A8A", size=10)
        section_font = Font(bold=True, color=ORANGE, size=11)
        m_lbl_font   = Font(bold=True, color="3A3A3A")
        m_val_font   = Font(bold=True, color="1F1F1F")
        lite_fill    = PatternFill("solid", fgColor="F5F5F5")
        right_al     = Alignment(horizontal="right", vertical="center")
        left_al      = Alignment(horizontal="left", vertical="center")

        # Banner + subtitle (full-width bar; span wide so the long title never clips)
        ws.merge_cells("A1:L1")
        for col in range(1, 13):
            ws.cell(1, col).fill = hdr_fill
        b = ws.cell(1, 1, "FireSense — Ringkasan Hasil Pemantauan")
        b.font, b.alignment = banner_font, Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 42
        ws.merge_cells("A2:L2")
        exported = datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M")
        s = ws.cell(2, 1, f"Pemantauan Firewall DQN   •   diekspor {exported}   •   {len(d_rows)} jendela")
        s.font, s.alignment = sub_font, Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18

        # metrics
        def num_col(name, cast=float):
            out = []
            if name in idx:
                for r in d_rows:
                    try: out.append(cast(r[idx[name]]))
                    except (ValueError, IndexError): pass
            return out
        lat = num_col("latency_ms", float)
        blk = num_col("blocklist_size", lambda x: int(float(x)))
        dist = collections.Counter(
            r[idx["action_name"]] for r in d_rows if "action_name" in idx and idx["action_name"] < len(r))
        def pretty_ts(v): return v.replace("T", " ") if isinstance(v, str) else v
        t_first = pretty_ts(d_rows[0][idx["timestamp"]]) if d_rows and "timestamp" in idx else "-"
        t_last  = pretty_ts(d_rows[-1][idx["timestamp"]]) if d_rows and "timestamp" in idx else "-"

        FMT_INT, FMT_MS = "#,##0", '#,##0.00" ms"'
        metrics = [
            ("Total jendela",       len(d_rows),                              FMT_INT),
            ("Waktu mulai",         t_first,                                  None),
            ("Waktu selesai",       t_last,                                   None),
            ("Latensi rata-rata",   round(sum(lat)/len(lat), 2) if lat else 0, FMT_MS),
            ("Latensi maksimum",    round(max(lat), 2) if lat else 0,          FMT_MS),
            ("Blocklist akhir",     blk[-1] if blk else 0,                     FMT_INT),
            ("Blocklist maksimum",  max(blk) if blk else 0,                    FMT_INT),
        ]
        ws.cell(4, 1, "IKHTISAR").font = section_font
        row = 5
        for label, value, fmt in metrics:
            lc = ws.cell(row, 1, label)
            lc.font, lc.fill, lc.alignment, lc.border = m_lbl_font, lite_fill, left_al, border
            vc = ws.cell(row, 2, self._to_num(value) if isinstance(value, str) else value)
            vc.font, vc.alignment, vc.border = m_val_font, right_al, border
            if fmt:
                vc.number_format = fmt
            row += 1

        # Distribusi aksi (tabel + data bar)
        row += 1
        ws.cell(row, 1, "DISTRIBUSI AKSI").font = section_font
        row += 1
        for col, txt in ((1, "Aksi"), (2, "Jumlah")):
            c = ws.cell(row, col, txt)
            c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, center, border
        first_data = row + 1
        for name in ("maintain", "block_ip", "tighten", "rollback"):
            row += 1
            ac = ws.cell(row, 1, name); ac.border, ac.alignment = border, left_al
            nc = ws.cell(row, 2, dist.get(name, 0)); nc.border, nc.alignment = border, center
            nc.number_format = FMT_INT
        last_data = row
        ws.conditional_formatting.add(
            f"B{first_data}:B{last_data}",
            DataBarRule(start_type="num", start_value=0, end_type="max",
                        color="F6B89B", showValue=True))

        # grafik batang tertanam (warna brand)
        chart = BarChart()
        chart.type, chart.title, chart.legend = "col", "Distribusi Aksi (Jumlah Jendela)", None
        chart.height, chart.width = 6.8, 12.5
        chart.add_data(Reference(ws, min_col=2, min_row=first_data - 1, max_row=last_data),
                       titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=first_data, max_row=last_data))
        chart.y_axis.majorGridlines = None
        # openpyxl gotcha: tanpa delete=False, Excel menyembunyikan label sumbu
        # meski kategori sudah di-set → batang tampil tanpa keterangan aksi.
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.majorTickMark = "out"
        # Tanpa judul sumbu terpisah: pada grafik kecil, judul sumbu ("Aksi"/
        # "Jumlah Jendela") mudah bertumpuk dgn label kategori & angka. Label
        # kategori sudah menyebut aksi, dan makna sumbu-Y dipindah ke judul grafik.
        try:
            chart.series[0].graphicalProperties.solidFill = ORANGE
        except Exception:
            pass
        ws.add_chart(chart, "D4")

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 2

        # ── Sheet 2: Keputusan ────────────────────────────────────────────────
        ws2 = wb.create_sheet("Keputusan")
        ws2.sheet_view.showGridLines = False
        write_table(ws2, d_header, d_rows)

        # ── Sheet 3: Kamus Data ───────────────────────────────────────────────
        if dict_csv and os.path.exists(dict_csv):
            dd = read_csv(dict_csv)
            if dd:
                ws3 = wb.create_sheet("Kamus Data")
                ws3.sheet_view.showGridLines = False
                write_table(ws3, dd[0], dd[1:], wrap_last=True)

        wb.save(dest_xlsx)

    def _export_results(self):
        """Export monitoring results as ONE styled Excel workbook inside ONE
        folder (with raw traffic snapshots), or just open the results folder."""
        results_dir = self._results_dir()
        src = os.path.join(results_dir, "window_decisions.csv")
        dict_src = os.path.join(results_dir, "window_decisions.dictionary.csv")
        if not os.path.exists(src):
            QMessageBox.information(
                self, "Belum ada hasil",
                "Berkas hasil pemantauan (window_decisions.csv) belum ada.\n"
                "Jalankan pemantauan minimal satu jendela terlebih dahulu.")
            return

        try:
            with open(src, "r", encoding="utf-8") as f:
                n_win = max(0, sum(1 for _ in f) - 1)
        except Exception:
            n_win = 0

        choice = QMessageBox.question(
            self, "Hasil Pemantauan",
            f"Tersedia {n_win} jendela.\n\n"
            "Pilih Ya untuk EKSPOR ke satu folder berisi buku kerja Excel "
            "(Ringkasan + Keputusan + Kamus Data), atau Tidak untuk BUKA FOLDER hasil.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            | QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Yes,
        )
        if choice == QMessageBox.StandardButton.Cancel:
            return
        if choice == QMessageBox.StandardButton.No:
            self._open_folder(results_dir)
            return

        parent = QFileDialog.getExistingDirectory(
            self, "Pilih folder tujuan ekspor", os.path.expanduser("~"))
        if not parent:
            return

        stamp   = time.strftime("%Y-%m-%d %H.%M.%S")
        out_dir = os.path.join(parent, f"FireSense — Hasil Pemantauan {stamp}")
        try:
            os.makedirs(out_dir, exist_ok=True)
            xlsx_path = os.path.join(out_dir, f"FireSense - Hasil Pemantauan {stamp}.xlsx")
            made_xlsx = False
            try:
                self._build_results_xlsx(src, dict_src, xlsx_path)
                made_xlsx = True
            except Exception as e:
                # openpyxl missing or failed → keep it usable: drop the raw CSVs in
                self.p_logs.append(f"[WARNING] Gagal membuat Excel ({e}); memakai CSV.")
                shutil.copyfile(src, os.path.join(out_dir, "window_decisions.csv"))
                if os.path.exists(dict_src):
                    shutil.copyfile(
                        dict_src, os.path.join(out_dir, "window_decisions.dictionary.csv"))

            # raw per-window snapshots go INTO the same export folder
            snap = os.path.join(results_dir, "traffic_snapshot")
            if os.path.isdir(snap):
                shutil.copytree(snap, os.path.join(out_dir, "traffic_snapshot"),
                                dirs_exist_ok=True)

            self.p_logs.append(
                f"[FireSense] Hasil pemantauan diekspor ke {out_dir} ({n_win} jendela).")
            doc = os.path.basename(xlsx_path) if made_xlsx else "window_decisions.csv"
            QMessageBox.information(
                self, "Ekspor selesai",
                f"Hasil pemantauan ({n_win} jendela) disimpan di satu folder:\n{out_dir}\n\n"
                f"Cukup buka: {doc}")
            self._open_folder(out_dir)
        except Exception as e:
            QMessageBox.warning(self, "Ekspor gagal", str(e))
            self.p_logs.append(f"[ERROR] Ekspor hasil gagal: {e}")

    def _open_folder(self, path):
        """Open a folder in the OS file manager (cross-platform)."""
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)                       # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                import subprocess; subprocess.Popen(["open", path])
            else:
                import subprocess; subprocess.Popen(["xdg-open", path])
            self.p_logs.append(f"[FireSense] Membuka folder hasil: {path}")
        except Exception as e:
            QMessageBox.information(
                self, "Folder hasil", f"Hasil pemantauan ada di:\n{path}\n\n({e})")

    def closeEvent(self, e):
        if self._worker and self._worker.isRunning():
            self._worker.stop(); self._worker.wait(3000)
        e.accept()
